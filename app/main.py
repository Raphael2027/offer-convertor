import os
import argparse
from datetime import datetime
import re
import unicodedata
import glob

import pandas as pd
import yaml
from rapidfuzz import process, fuzz
from openpyxl import load_workbook

def detect_header_row(raw, max_rows=50):
    return 0


# ---------- Helpers ----------

def load_yaml(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def normalize(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    # Umlaute & Akzente entfernen
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    # Klammerinhalte entfernen (z.B. "Preis (CHF)")
    s = re.sub(r"\(.*?\)", " ", s)
    # Sonderzeichen vereinheitlichen
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def is_mostly_numeric(series, sample=30):
    vals = [v for v in series.dropna().head(sample).tolist() if str(v).strip() != ""]
    if not vals:
        return False
    ok = 0
    for v in vals:
        try:
            float(str(v).replace(",", "."))
            ok += 1
        except:
            pass
    return ok / len(vals) >= 0.8

def detect_currency_from_columns(columns):
    joined = " ".join([str(c) for c in columns]).upper()
    if "USD" in joined:
        return "USD"
    if "EUR" in joined:
        return "EUR"
    if "CHF" in joined:
        return "CHF"
    return "EUR"  # Default

def looks_like_ean(series, sample=30):
    vals = [v for v in series.dropna().head(sample).tolist() if str(v).strip() != ""]
    if not vals:
        return False
    ok = 0
    for v in vals:
        s = str(v).strip()
        if s.endswith(".0"):
            s = s[:-2]
        digits = "".join(ch for ch in s if ch.isdigit())
        if len(digits) in (8, 12, 13, 14) and digits == s:
            ok += 1
    return ok / len(vals) >= 0.7

def fuzzy_map(input_cols, target_cols, synonyms, threshold=85):
    mapped = {}
    unmapped = []

    candidates = {t: [t] + synonyms.get(t, []) for t in target_cols}

    for col in input_cols:
        col_n = normalize(col)
        best_t = None
        best_s = -1

        for t, cands in candidates.items():
            match, score, _ = process.extractOne(
                col_n,
                [normalize(c) for c in cands],
                scorer=fuzz.WRatio
            )
            if score > best_s:
                best_s = score
                best_t = t

        if best_s >= threshold and best_t not in mapped.values():
            mapped[col] = best_t
        else:
            unmapped.append(col)

    return mapped, unmapped

def load_all_supplier_configs(base):
    configs = []
    pattern = os.path.join(base, "configs", "suppliers", "*.yml")
    for path in glob.glob(pattern):
        cfg = load_yaml(path)
        cfg["_path"] = path
        configs.append(cfg)
    return configs


def choose_supplier_config(df_columns, configs):
    norm_cols = [normalize(c) for c in df_columns]

    best_cfg = None
    best_score = -1

    for cfg in configs:
        score = 0

        for src in cfg.get("mapping", {}).keys():
            if normalize(src) in norm_cols:
                score += 3

        for syns in cfg.get("synonyms", {}).values():
            for s in syns:
                if normalize(s) in norm_cols:
                    score += 1

        if score > best_score:
            best_score = score
            best_cfg = cfg

    return best_cfg

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        required=True,
        help="Path to supplier Excel file"
    )
    args = parser.parse_args()
    unmapped = []


    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # -------------------------------------------------
    # Load schema (Zielstruktur)
    # -------------------------------------------------
    schema = load_yaml(os.path.join(base, "configs", "schema.yml"))
    target_cols = [c["name"] for c in schema["columns"]]
    required_cols = [c["name"] for c in schema["columns"] if c.get("required")]

    # -------------------------------------------------
    # Load Excel RAW (ohne Header, Header automatisch finden)
    # -------------------------------------------------
    raw = pd.read_excel(args.input, header=None, dtype=object)
    header_row = detect_header_row(raw)
    df = pd.read_excel(args.input, header=header_row, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]

    # ----------------------------
    # Auto Supplier Detection
    # ----------------------------
    supplier_cfgs = load_all_supplier_configs(base)
    supplier_cfg = choose_supplier_config(df.columns, supplier_cfgs)

    if supplier_cfg is None:
        raise ValueError("❌ Kein passendes Supplier-Mapping gefunden")


    # -------------------------------------------------
    # Build standardized output dataframe
    # -------------------------------------------------
    out = pd.DataFrame()

    out["supplier"] = supplier_cfg.get("supplier_name", "UNKNOWN")

    for col in target_cols:
        if col == "supplier":
            continue
        out[col] = df[col] if col in df.columns else None

    out = out[target_cols]

    # -------------------------------------------------
    # Fallbacks & Defaults
    # -------------------------------------------------
    if "description" in out.columns and "product_name" in out.columns:
        out["description"] = out["description"].fillna(out["product_name"])

    if "currency" in out.columns and out["currency"].isna().all():
        out["currency"] = detect_currency_from_columns(df.columns)

    # -------------------------------------------------
    # Validation Sheet
    # -------------------------------------------------
    messages = []
    missing = [c for c in required_cols if out[c].isna().all()]

    if missing:
        messages.append(
            "ERROR: missing required columns: " + ", ".join(missing)
        )

    if unmapped:
        messages.append(
            "WARN: unmapped input columns: " + ", ".join(unmapped)
        )

    messages.append(f"INFO: header row detected: {header_row}")
    messages.append(f"INFO: rows exported: {len(out)}")

    # -------------------------------------------------
    # Write Excel Output (IMMER EIN FILE)
    # -------------------------------------------------
    os.makedirs(os.path.join(base, "output"), exist_ok=True)

    out_path = os.path.join(base, "output", "normalized.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Data", index=False)
        pd.DataFrame({"log": messages}).to_excel(
            writer, sheet_name="Validation", index=False
        )

    print(f"✅ Output written to: {out_path}")

if __name__ == "__main__":
    main()


# ---------- Header Detection ----------

def detect_header_row(df, max_rows=50):
    best_row = 0
    best_score = -1

    for i in range(min(len(df), max_rows)):
        row = df.iloc[i].tolist()
        strings = sum(isinstance(x, str) and x.strip() != "" for x in row)
        non_empty = sum(x is not None and str(x).strip() != "" for x in row)
        score = strings * 2 + non_empty

        if score > best_score:
            best_score = score
            best_row = i

    return best_row


# ---------- Column Mapping ----------

def fuzzy_map(df, input_cols, target_cols, synonyms, threshold=88):
    mapped = {}
    unmapped = []

    candidates = {t: [t] + synonyms.get(t, []) for t in target_cols}

    for col in input_cols:
        col_n = normalize(col)

        # Listenpreise ignorieren (sonst landet MSRP als unit_price)
        if "msrp" in col_n:
            unmapped.append(col)
            continue

        # Header-Hints (Boost)
        boost_unit_price = any(x in col_n for x in ["cfr", "fob", "exw", "price", "cost"])
        boost_qty = any(x in col_n for x in ["qty", "quantity", "menge", "anzahl", "total qty"])

        best_t, best_s = None, -1

        for t, cands in candidates.items():
            _, score, _ = process.extractOne(
                col_n,
                [normalize(c) for c in cands],
                scorer=fuzz.WRatio
            )

            if t == "unit_price" and boost_unit_price:
                score += 7
            if t == "quantity" and boost_qty:
                score += 7

            if score > best_s:
                best_s = score
                best_t = t

        # Zu unsicher -> nicht mappen
        if best_s < threshold or best_t is None:
            unmapped.append(col)
            continue

        # Safety checks auf Basis echter Spaltendaten
        series = df[col] if col in df.columns else pd.Series([], dtype=object)

        # currency darf nicht numerisch sein
        if best_t == "currency" and is_mostly_numeric(series):
            unmapped.append(col)
            continue

        # ean muss wie ean aussehen
        if best_t == "ean" and not looks_like_ean(series):
            unmapped.append(col)
            continue

        # Zielspalte nicht doppelt vergeben
        if best_t in mapped.values():
            unmapped.append(col)
            continue

        mapped[col] = best_t

    return mapped, unmapped

def detect_header_row(raw, max_rows=50):
    """
    Ermittelt die Header-Zeile in einem rohen Excel-Import.
    Heuristik: erste Zeile mit genügend nicht-leeren String-Werten.
    """
    for i in range(min(len(raw), max_rows)):
        row = raw.iloc[i]
        non_empty = row.dropna()
        string_cells = non_empty[non_empty.map(lambda x: isinstance(x, str))]

        if len(string_cells) >= len(row) * 0.5:
            return i

    # Fallback
    return 0

# ---------- Main ----------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Supplier Excel file")
    parser.add_argument("--supplier", default="generic")
    args = parser.parse_args()

    base = os.path.dirname(os.path.dirname(__file__))

    schema = load_yaml(os.path.join(base, "configs", "schema.yml"))
    target_cols = [c["name"] for c in schema["columns"]]
    required_cols = {c["name"] for c in schema["columns"] if c["required"]}

    # Load raw without header
    raw = pd.read_excel(args.input, header=None, dtype=object)
    header_row = detect_header_row(raw, max_rows=50)

    # Reload with header
    df = pd.read_excel(args.input, header=header_row, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]
    supplier_cfgs = load_all_supplier_configs(base)
    supplier_cfg = choose_supplier_config(df.columns, supplier_cfgs)
    if supplier_cfg is None:
        raise ValueError("❌ Kein passendes Supplier-Mapping gefunden")

    # Explicit mapping
    rename = {}
    for k, v in supplier_cfg.get("mapping", {}).items():
        for col in df.columns:
            if normalize(col) == normalize(k):
                rename[col] = v

    df = df.rename(columns=rename)

    # Fuzzy Mapping (Synonyme & ähnliche Begriffe)
    remaining_cols = [c for c in df.columns if c not in target_cols]
    remaining_df = df[remaining_cols]

    fuzzy, unmapped = fuzzy_map(
        remaining_df,
        remaining_df.columns,
        target_cols,
        supplier_cfg.get("synonyms", {}),
    )
  
    df = df.rename(columns=fuzzy)


    # Build standardized df
    out = pd.DataFrame()
    out["supplier"] = supplier_cfg.get("supplier_name", "UNKNOWN")

    for col in target_cols:
        if col == "supplier":
            continue
        out[col] = df[col] if col in df.columns else None

    out = out[target_cols]

    # Fallback: description = product_name
    if "description" in out.columns and "product_name" in out.columns:
        out["description"] = out["description"].fillna(out["product_name"])

    # Fill currency if missing (from input header, otherwise default)
    if "currency" in out.columns and out["currency"].isna().all():
        out["currency"] = detect_currency_from_columns(df.columns)

    # Validation
    messages = []
    missing = [c for c in required_cols if out[c].isna().all()]
    if missing:
        messages.append("ERROR: missing required columns: " + ", ".join(missing))
    if unmapped:
        messages.append("WARN: unmapped input columns: " + ", ".join(map(str, unmapped)))
    messages.append(f"INFO: header row detected: {header_row}")
    messages.append(f"INFO: rows exported: {len(out)}")

    # Write to template
    wb = load_workbook(os.path.join(base, "templates", "template.xlsx"))
    ws = wb[schema["output_sheet"]]

    start = schema["start_row"]
    header_row_out = start - 1

    for i, col in enumerate(out.columns, start=1):
        ws.cell(row=header_row_out, column=i, value=col)

    for r, row in enumerate(out.itertuples(index=False), start=start):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    # Validation sheet
    if "Validation" in wb.sheetnames:
        vws = wb["Validation"]
        vws.delete_rows(1, vws.max_row)
    else:
        vws = wb.create_sheet("Validation")

    for i, msg in enumerate(messages, start=1):
        vws.cell(row=i, column=1, value=msg)

    os.makedirs(os.path.join(base, "output"), exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(base, "output", f"normalized_{ts}.xlsx")
    wb.save(out_path)

    print(f"✔ Output written to: {out_path}")


if __name__ == "__main__":
    main()
